"""
Scraper: AHDB livestock prices.
- Pig SPP (Standard Pig Price): weekly Excel download
- Milk farmgate price: monthly Excel download
- Beef deadweight: weekly Excel (found via page scrape)
- Egg packer-to-producer price: weekly Excel (found via page scrape)
- Poultry: Excel if available, else commentary link
Outputs: data/ahdb_livestock.json
"""

import logging
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
import xlrd
from bs4 import BeautifulSoup

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from scrapers.base import archive_current, download_excel, get, load_previous, save_data

logger = logging.getLogger("field_notes.ahdb_livestock")

# .xls format (use xlrd)
PIG_SPP_URL = (
    "https://projectblue.blob.core.windows.net/media/Default/Market%20Intelligence/"
    "pork/Prices/Deadweight%20pig%20prices/GB%20SPP%20Full%20Report.xls"
)
MILK_URL = (
    "https://projectblue.blob.core.windows.net/media/Default/Market%20Intelligence/"
    "dairy/Files/Prices/Farmgate%20milk%20prices/UK%20GB%20and%20NI%20Farmgate%20prices.xlsx"
)
BEEF_PAGE_URL = "https://ahdb.org.uk/beef-and-lamb/prices-and-markets"
EGG_AHDB_PAGE_URL = "https://ahdb.org.uk/egg-prices"
POULTRY_PAGE_URL = "https://ahdb.org.uk/poultry"


def _latest_two_rows(ws, date_col=0, price_col=1):
    """Return (latest_price, latest_date_str, prev_price) from the last two data rows."""
    data_rows = []
    for row in ws.iter_rows(values_only=True):
        val_date = row[date_col] if len(row) > date_col else None
        val_price = row[price_col] if len(row) > price_col else None
        if val_date and val_price is not None:
            try:
                price = float(val_price)
                if isinstance(val_date, datetime):
                    date_str = val_date.strftime("%Y-%m-%d")
                else:
                    date_str = str(val_date).strip()
                data_rows.append((date_str, price))
            except (ValueError, TypeError):
                continue
    if not data_rows:
        return None, None, None
    latest_date, latest = data_rows[-1]
    prev = data_rows[-2][1] if len(data_rows) >= 2 else None
    return latest, latest_date, prev


def _scrape_pig_spp() -> dict:
    """
    Parse GB SPP Full Report.xls.
    Row 2 (0-based): ['Category: ', 'Standard Pig Price', 'Week ended: 14/03/2026']
    Row 16: ['GB SPP (UK Spec) ', price, change_on_week, ...]
    """
    logger.info("Downloading pig SPP Excel")
    try:
        buf = download_excel(PIG_SPP_URL)
        wb = xlrd.open_workbook(file_contents=buf.read())
        ws = wb.sheets()[0]

        # Extract week-ending — scan all cells in first few rows
        week_ending = None
        for ri in range(min(5, ws.nrows)):
            for cell in ws.row_values(ri):
                if cell and "week end" in str(cell).lower():
                    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", str(cell))
                    if m:
                        week_ending = f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
                    break
            if week_ending:
                break

        # Find GB SPP (UK Spec) row — price at col15, week-on-week change at col22
        price = None
        change = None
        for i in range(ws.nrows):
            row = ws.row_values(i)
            if row and "gb spp (uk spec)" in str(row[0]).lower():
                try:
                    # price col and change col found at col15 and col22 empirically
                    non_empty = [(j, v) for j, v in enumerate(row) if v not in ('', None, 0, 0.0)]
                    if len(non_empty) >= 2:
                        price = float(str(non_empty[1][1]).replace(",", ""))
                    if len(non_empty) >= 3:
                        change = float(str(non_empty[2][1]).replace(",", ""))
                except (ValueError, TypeError):
                    pass
                break

        if price is None:
            return {"error": "No GB SPP (UK Spec) row found in pig SPP Excel"}

        previous = load_previous("ahdb_livestock.json")
        prev_week = previous.get("pig_prices", {}).get("price")
        effective_change = change if change is not None else (
            round(price - prev_week, 2) if prev_week else None
        )
        return {
            "series": "GB Standard Pig Price (UK Spec)",
            "unit": "p/kg deadweight",
            "week_ending": week_ending,
            "price": round(price, 2),
            "prev_week_price": prev_week,
            "change": round(effective_change, 2) if effective_change is not None else None,
        }
    except Exception as exc:
        logger.error("Pig SPP scrape failed: %s", exc)
        return {"error": str(exc)}


def _scrape_milk() -> dict:
    """
    Parse UK GB and NI Farmgate prices.xlsx.
    Sheet 'UK average farmgate price': col1=date, col2=price (ppl) — 0-indexed, first col is blank/header.
    """
    logger.info("Downloading milk farmgate Excel")
    try:
        buf = download_excel(MILK_URL)
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb["UK average farmgate price"]
        price, date_str, prev = _latest_two_rows(ws, date_col=1, price_col=2)
        if price is None:
            return {"error": "No milk price data found"}
        change = round(price - prev, 2) if prev is not None else None
        return {
            "series": "UK Farmgate Milk Price",
            "unit": "ppl",
            "period": date_str,
            "price": round(price, 2),
            "prev_period_price": round(prev, 2) if prev else None,
            "change": change,
        }
    except Exception as exc:
        logger.error("Milk price scrape failed: %s", exc)
        return {"error": str(exc)}


def _find_excel_url(page_url: str, keywords: list) -> Optional[str]:
    """Scrape a page to find the first Excel download URL containing any keyword."""
    try:
        resp = get(page_url)
        soup = BeautifulSoup(resp.text, "lxml")
        for a in soup.find_all("a", href=True):
            href = str(a.get("href", ""))
            link_text = a.get_text(strip=True).lower()
            if ".xlsx" not in href.lower() and ".xls" not in href.lower():
                continue
            if any(kw.lower() in link_text or kw.lower() in href.lower() for kw in keywords):
                if href.startswith("http"):
                    return href
                if href.startswith("/"):
                    return "https://ahdb.org.uk" + href
    except Exception as exc:
        logger.warning("Could not find Excel on %s: %s", page_url, exc)
    return None


def _scrape_beef() -> dict:
    """Scrape AHDB GB cattle deadweight prices (weekly Excel)."""
    logger.info("Fetching AHDB beef deadweight prices")
    try:
        excel_url = _find_excel_url(BEEF_PAGE_URL, ["deadweight", "cattle", "beef", "weekly"])
        if not excel_url:
            return {"error": "Could not find beef deadweight Excel on AHDB page"}

        buf = download_excel(excel_url)
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)

        # Find the most relevant sheet
        target_ws = None
        for name in wb.sheetnames:
            name_lower = name.lower()
            if any(kw in name_lower for kw in ["deadweight", "cattle", "gb", "all"]):
                target_ws = wb[name]
                break
        if target_ws is None:
            target_ws = wb.active

        price, date_str, prev = _latest_two_rows(target_ws, date_col=0, price_col=1)
        if price is None:
            return {"error": "No beef price data found in Excel"}

        previous = load_previous("ahdb_livestock.json")
        prev_week = previous.get("beef_prices", {}).get("price")
        change = round(price - (prev if prev is not None else prev_week), 2) if (prev is not None or prev_week is not None) else None

        return {
            "series": "GB Cattle Deadweight",
            "unit": "p/kg dwt",
            "week_ending": date_str,
            "price": round(price, 2),
            "prev_week_price": prev_week,
            "change": change,
        }
    except Exception as exc:
        logger.error("Beef price scrape failed: %s", exc)
        return {"error": str(exc)}


def _scrape_eggs() -> dict:
    """Scrape AHDB UK egg packer-to-producer prices (weekly Excel)."""
    logger.info("Fetching AHDB egg prices")
    try:
        excel_url = _find_excel_url(
            EGG_AHDB_PAGE_URL,
            ["egg", "packer", "producer", "weekly", "price"],
        )
        if not excel_url:
            return {"error": "Could not find egg prices Excel on AHDB page"}

        buf = download_excel(excel_url)
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)

        target_ws = None
        for name in wb.sheetnames:
            name_lower = name.lower()
            if any(kw in name_lower for kw in ["egg", "packer", "producer", "price"]):
                target_ws = wb[name]
                break
        if target_ws is None:
            target_ws = wb.active

        price, date_str, prev = _latest_two_rows(target_ws, date_col=0, price_col=1)
        if price is None:
            return {"error": "No egg price data found in Excel"}

        previous = load_previous("ahdb_livestock.json")
        prev_week = previous.get("egg_prices", {}).get("price")
        change = round(price - (prev if prev is not None else prev_week), 2) if (prev is not None or prev_week is not None) else None

        return {
            "series": "UK Egg Packer-to-Producer",
            "unit": "p/doz",
            "week_ending": date_str,
            "price": round(price, 2),
            "prev_week_price": prev_week,
            "change": change,
        }
    except Exception as exc:
        logger.error("Egg price scrape failed: %s", exc)
        return {"error": str(exc)}


def _scrape_poultry() -> dict:
    """Attempt to get AHDB poultry market prices; fall back to commentary link."""
    logger.info("Fetching AHDB poultry data")
    try:
        excel_url = _find_excel_url(
            POULTRY_PAGE_URL,
            ["poultry", "chicken", "broiler", "turkey", "price", "weekly"],
        )
        if not excel_url:
            return {"note": "See ahdb.org.uk/poultry for poultry market commentary.", "source_url": POULTRY_PAGE_URL}

        buf = download_excel(excel_url)
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        target_ws = wb.active
        price, date_str, prev = _latest_two_rows(target_ws, date_col=0, price_col=1)
        if price is None:
            return {"note": "See ahdb.org.uk/poultry for poultry market commentary.", "source_url": POULTRY_PAGE_URL}

        previous = load_previous("ahdb_livestock.json")
        prev_week = previous.get("poultry_prices", {}).get("price")
        change = round(price - (prev if prev is not None else prev_week), 2) if (prev is not None or prev_week is not None) else None

        return {
            "series": "UK Poultry",
            "unit": "p/kg",
            "week_ending": date_str,
            "price": round(price, 2),
            "prev_week_price": prev_week,
            "change": change,
        }
    except Exception as exc:
        logger.error("Poultry price scrape failed: %s", exc)
        return {"note": "See ahdb.org.uk/poultry for poultry market commentary.", "source_url": POULTRY_PAGE_URL}


def scrape() -> dict:
    pig = _scrape_pig_spp()
    milk = _scrape_milk()
    eggs = _scrape_eggs()
    beef = _scrape_beef()
    poultry = _scrape_poultry()

    logger.info(
        "Livestock prices: pig=%s p/kg, milk=%s ppl, beef=%s p/kg",
        pig.get("price", "ERR"),
        milk.get("price", "ERR"),
        beef.get("price", "ERR"),
    )
    return {
        "source": "AHDB",
        "pig_prices": pig,
        "milk_prices": milk,
        "egg_prices": eggs,
        "beef_prices": beef,
        "poultry_prices": poultry,
    }


if __name__ == "__main__":
    archive_current("ahdb_livestock.json")
    data = scrape()
    save_data("ahdb_livestock.json", data)
    print(f"Pig SPP: {data['pig_prices'].get('price', 'ERR')} p/kg dwt")
    print(f"Milk: {data['milk_prices'].get('price', 'ERR')} ppl")
    print(f"Beef: {data['beef_prices'].get('price', 'ERR')} p/kg dwt")
    print(f"Eggs: {data['egg_prices'].get('price', 'ERR')} p/doz")
